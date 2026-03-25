"""
Módulo de procesamiento de datos para evidencias de cobranzas
Maneja la sanitización de campos y generación de archivos de evidencias
"""
import pandas as pd
import numpy as np
import os
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import numbers
from typing import Dict, List, Tuple, Optional


class ProcesadorDatos:
    """Procesador de datos para generación de evidencias de gestión"""
    
    def __init__(self, funcion_log=None):
        """
        Inicializa el procesador de datos
        
        Args:
            funcion_log: Función para enviar mensajes de log a la interfaz
        """
        self.funcion_log = funcion_log
        
        # Mapeo de nombres de campos para sanitización
        self.mapeo_campos = {
            'cuenta': ['cuenta', 'CUENTA', 'Cuenta'],
            'nombre': ['nombre', 'NOMBRE', 'nombres', 'NOMBRES', 'contacto', 'CONTACTO', 
                      'nombre completo', 'NOMBRE COMPLETO', 'nombre_completo', 'NOMBRE_COMPLETO'],
            'dni': ['dni', 'DNI', 'documento', 'DOCUMENTO', 'Dni', 'Documento'],
            'gestion_efectiva': ['gestion efectiva', 'GESTION EFECTIVA', 'gestión efectiva', 
                                'GESTIÓN EFECTIVA', 'gestion_efectiva', 'GESTION_EFECTIVA'],
            'telefono': ['telefono', 'TELEFONO', 'teléfono', 'TELÉFONO', 'celular', 
                        'CELULAR', 'Telefono', 'Celular'],
            'tipo_gestion': ['tipo de gestion', 'TIPO DE GESTION', 'tipo_gestion', 'TIPO_GESTION',
                           'tipo de gestión', 'TIPO DE GESTIÓN'],
            'numero_credito': ['numero de credito', 'NUMERO DE CREDITO', 'número de crédito',
                             'NÚMERO DE CRÉDITO', 'numero_credito', 'NUMERO_CREDITO'],
            'ruta': ['ruta', 'RUTA', 'Ruta'],
            'nombre_completo_audio': ['nombre_completo', 'NOMBRE_COMPLETO', 'nombre completo']
        }
    
    def registrar_log(self, mensaje: str):
        """Envía un mensaje de log a la interfaz"""
        if self.funcion_log:
            self.funcion_log(mensaje)
            
    def limpiar_id(self, valor):
        """
        Limpia IDs numéricos (DNI, Teléfono, Cuenta) de forma segura.
        Preserva la precisión total evitando conversiones a float.
        """
        if pd.isna(valor) or valor == '':
            return ""
        
        # Convertir a string de forma segura
        val_s = str(valor).strip()
        
        # Si tiene .0 al final (típico de floats en Excel), lo quitamos SIN usar float()
        if val_s.endswith('.0'):
            val_s = val_s[:-2]
            
        return val_s
    
    def guardar_excel_formateado(self, df: pd.DataFrame, ruta_excel: Path):
        """
        Guarda un DataFrame a Excel con formato de texto para campos numéricos
        y sin valores NaN (se muestran como celdas vacías)
        """
        # Crear copia para no modificar el original
        df_formateado = df.copy()
        
        # Reemplazar NaN con cadena vacía
        df_formateado = df_formateado.fillna('')
        
        # Columnas que deben ser texto puro
        columnas_numericas = ['cuenta', 'telefono', 'celular', 'dni', 'documento',
                           'numero_credito', 'CUENTA', 'TELEFONO', 'CELULAR', 'DNI',
                           'DOCUMENTO', 'NUMERO DE CREDITO', 'numero de credito']
        
        # Asegurar que todas las columnas ID sean strings limpios
        for col in df_formateado.columns:
            col_minusc = col.lower().strip()
            if col in columnas_numericas or col_minusc in [c.lower() for c in columnas_numericas]:
                df_formateado[col] = df_formateado[col].apply(self.limpiar_id)
        
        # Guardar el archivo Excel
        df_formateado.to_excel(ruta_excel, index=False, engine='openpyxl')
        
        # Aplicar formato de texto en openpyxl para que Excel lo reconozca como string
        libro = load_workbook(ruta_excel)
        
        # Resetear metadatos internos del Excel
        ahora = datetime.now()
        libro.properties.created = ahora
        libro.properties.modified = ahora
        libro.properties.lastModifiedBy = "Sistema de Evidencias"
        
        hoja = libro.active
        
        fila_encabezado = list(hoja.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        indices_col_numericas = []
        for idx, nombre_col in enumerate(fila_encabezado, 1):
            if nombre_col:
                col_minusc = str(nombre_col).lower().strip()
                if nombre_col in columnas_numericas or col_minusc in [c.lower() for c in columnas_numericas]:
                    indices_col_numericas.append(idx)
        
        for col_idx in indices_col_numericas:
            for fila in range(2, hoja.max_row + 1):
                celda = hoja.cell(row=fila, column=col_idx)
                celda.number_format = numbers.FORMAT_TEXT
        
        libro.save(ruta_excel)
        
        # Asegurar que la fecha de modificación del archivo en el sistema sea la actual
        os.utime(ruta_excel, None)
    
    def sanitizar_dataframe(self, df: pd.DataFrame, omitir_consolidados: bool = False) -> pd.DataFrame:
        """
        Sanitiza los nombres de columnas de un DataFrame
        
        Args:
            df: DataFrame a sanitizar
            omitir_consolidados: Si es True, no sanitiza (para consolidados.xlsx)
            
        Returns:
            DataFrame con columnas sanitizadas
        """
        if omitir_consolidados:
            # Solo quitar espacios en blanco de valores, no cambiar nombres de columnas
            df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
            return df
        
        df_copia = df.copy()
        
        # Renombrar columnas según el mapeo
        renombrar_columna = {}
        for nombre_estandar, variaciones in self.mapeo_campos.items():
            for col in df_copia.columns:
                if col.strip() in variaciones:
                    renombrar_columna[col] = nombre_estandar
                    break
        
        df_copia.rename(columns=renombrar_columna, inplace=True)
        
        # Quitar espacios en blanco de los valores y limpiar IDs
        for col in df_copia.columns:
            col_minusc = col.lower().strip()
            # Identificar columnas que deben ser tratadas como IDs limpios
            es_col_id = col in ['cuenta', 'dni', 'telefono', 'numero_credito', 'documento', 'celular'] or \
                         col_minusc in ['cuenta', 'dni', 'telefono', 'numero_credito', 'documento', 'celular']
            
            if es_col_id:
                df_copia[col] = df_copia[col].apply(self.limpiar_id)
            elif df_copia[col].dtype == 'object':
                df_copia[col] = df_copia[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        
        return df_copia
    
    def analizar_gestion_efectiva(self, gestion_str: str) -> List[str]:
        """
        Parsea el campo GESTION EFECTIVA separado por comas
        
        Args:
            gestion_str: String con gestiones separadas por coma
            
        Returns:
            Lista de gestiones (IVR, SMS, CALL, GRABACION CALL)
        """
        if pd.isna(gestion_str):
            return []
        
        gestiones = [g.strip().upper() for g in str(gestion_str).split(',')]
        
        # Normalizar GRABACION CALL a CALL
        gestiones = ['CALL' if 'CALL' in g else g for g in gestiones]
        
        return list(set(gestiones))  # Eliminar duplicados
    
    def crear_evidencia_ivr(self, datos_cliente: Dict, df_nuevos_datos: pd.DataFrame, 
                           carpeta_salida: Path, ruta_audio_ivr: str) -> Tuple[bool, List[str]]:
        """
        Crea archivos de evidencia IVR para un cliente
        
        Returns:
            Tuple (exito, archivos_creados)
        """
        archivos_creados = []
        
        try:
            cuenta = datos_cliente['cuenta']
            nombre = datos_cliente['nombre']
            
            # Copiar audio IVR (SIEMPRE se copia si el cliente tiene gestión IVR)
            nombre_archivo_audio = f"ivr_{nombre}.mp3"
            ruta_audio = carpeta_salida / nombre_archivo_audio
            
            # Eliminar si existe para asegurar nueva fecha de creación
            if ruta_audio.exists():
                ruta_audio.unlink()
                
            shutil.copy(ruta_audio_ivr, ruta_audio)
            os.utime(ruta_audio, None) # Forzar fecha actual
            archivos_creados.append(nombre_archivo_audio)
            
            # Filtrar en nuevos_datos por CUENTA y GESTION_EFECTIVA = IVR
            datos_ivr = df_nuevos_datos[
                (df_nuevos_datos['cuenta'] == cuenta) & 
                (df_nuevos_datos['gestion_efectiva'].str.contains('IVR', na=False))
            ].copy()
            
            if datos_ivr.empty:
                self.registrar_log(f"  ⚠️ No se encontraron registros IVR en nuevos_datos para {nombre} (audio IVR copiado)")
            else:
                # Agregar columna TIPO DE GESTION
                datos_ivr['TIPO DE GESTION'] = 'IVR'
                
                # Crear archivo Excel con formato de texto para campos numéricos
                nombre_archivo_excel = f"{nombre}_ivr.xlsx"
                ruta_excel = carpeta_salida / nombre_archivo_excel
                self.guardar_excel_formateado(datos_ivr, ruta_excel)
                archivos_creados.append(nombre_archivo_excel)
            
            return True, archivos_creados
            
        except Exception as e:
            self.registrar_log(f"  ❌ Error creando evidencia IVR: {str(e)}")
            return False, archivos_creados
    
    def crear_evidencia_sms(self, datos_cliente: Dict, df_sms: pd.DataFrame, 
                           carpeta_salida: Path) -> Tuple[bool, List[str]]:
        """
        Crea archivo de evidencia SMS para un cliente
        
        Returns:
            Tuple (exito, archivos_creados)
        """
        archivos_creados = []
        
        try:
            cuenta = datos_cliente['cuenta']
            nombre = datos_cliente['nombre']
            
            # Filtrar en sms.xlsx por NUMERO DE CREDITO
            datos_sms = df_sms[df_sms['numero_credito'] == cuenta].copy()
            
            if datos_sms.empty:
                self.registrar_log(f"  ⚠️ No se encontraron registros SMS para {nombre}")
                return False, archivos_creados
            
            # Crear archivo Excel con formato de texto para campos numéricos
            nombre_archivo_excel = f"SMS_{nombre}.xlsx"
            ruta_excel = carpeta_salida / nombre_archivo_excel
            self.guardar_excel_formateado(datos_sms, ruta_excel)
            archivos_creados.append(nombre_archivo_excel)
            
            return True, archivos_creados
            
        except Exception as e:
            self.registrar_log(f"  ❌ Error creando evidencia SMS: {str(e)}")
            return False, archivos_creados
    
    def crear_evidencia_call(self, datos_cliente: Dict, df_nuevos_datos: pd.DataFrame,
                            df_consolidados: Optional[pd.DataFrame], carpeta_salida: Path) -> Tuple[bool, List[str]]:
        """
        Crea archivos de evidencia CALL para un cliente
        
        Returns:
            Tuple (exito, archivos_creados)
        """
        archivos_creados = []
        
        try:
            cuenta = datos_cliente['cuenta']
            nombre = datos_cliente['nombre']
            dni = datos_cliente.get('dni', '')
            telefono = datos_cliente.get('telefono', '')
            
            # Filtrar en nuevos_datos por CUENTA y GESTION_EFECTIVA = CALL
            datos_call = df_nuevos_datos[
                (df_nuevos_datos['cuenta'] == cuenta) & 
                (df_nuevos_datos['gestion_efectiva'].str.contains('CALL', na=False))
            ].copy()
            
            if datos_call.empty:
                self.registrar_log(f"  ⚠️ No se encontraron registros CALL en nuevos_datos para {nombre}")
                return False, archivos_creados
            
            # Agregar columna TIPO DE GESTION
            datos_call['TIPO DE GESTION'] = 'CALL'
            
            # Crear archivo Excel con formato de texto para campos numéricos
            nombre_archivo_excel = f"{nombre}_gestiones.xlsx"
            ruta_excel = carpeta_salida / nombre_archivo_excel
            self.guardar_excel_formateado(datos_call, ruta_excel)
            archivos_creados.append(nombre_archivo_excel)
            
            # Buscar audio y transcripción en consolidados (OPCIONAL - solo si existe df_consolidados)
            if df_consolidados is not None:
                audio_encontrado = False
                fila_audio = None
                
                # Primero intentar buscar por DNI
                if dni:
                    dni_limpio = self.limpiar_id(dni)
                    fila_audio = df_consolidados[df_consolidados['dni'].apply(self.limpiar_id) == dni_limpio]
                    if not fila_audio.empty:
                        audio_encontrado = True
                
                # Si no se encontró por DNI, buscar por teléfono
                if not audio_encontrado and telefono:
                    tel_limpio = self.limpiar_id(telefono)
                    fila_audio = df_consolidados[df_consolidados['telefono'].apply(self.limpiar_id) == tel_limpio]
                    if not fila_audio.empty:
                        audio_encontrado = True
                
                if audio_encontrado and not fila_audio.empty:
                    # Obtener nombre_completo para buscar audio y transcripción
                    ruta = str(fila_audio.iloc[0]['ruta'])
                    audio_nombre_completo = str(fila_audio.iloc[0]['nombre_completo'])
                    
                    # 1. Copiar audio MP3
                    ruta_origen_audio = f"{ruta}/{audio_nombre_completo}.mp3"
                    
                    if os.path.exists(ruta_origen_audio):
                        nombre_archivo_audio = f"{nombre}_{cuenta}.mp3"
                        ruta_destino_audio = carpeta_salida / nombre_archivo_audio
                        
                        # Eliminar si existe
                        if ruta_destino_audio.exists():
                            ruta_destino_audio.unlink()
                            
                        shutil.copy(ruta_origen_audio, ruta_destino_audio)
                        os.utime(ruta_destino_audio, None)
                        archivos_creados.append(nombre_archivo_audio)
                    else:
                        self.registrar_log(f"  ⚠️ Audio no encontrado en: {ruta_origen_audio}")
                    
                    # 2. Buscar y copiar archivo de transcripción TXT
                    ruta_base_transcripcion = "E:/ProcesoAudios/2025/everyVerse/15-19/evidencias_general"
                    ruta_origen_transcripcion = f"{ruta_base_transcripcion}/{audio_nombre_completo}.txt"
                    
                    if os.path.exists(ruta_origen_transcripcion):
                        nombre_archivo_transcripcion = f"{nombre}_{cuenta}.txt"
                        ruta_destino_transcripcion = carpeta_salida / nombre_archivo_transcripcion
                        
                        # Eliminar si existe
                        if ruta_destino_transcripcion.exists():
                            ruta_destino_transcripcion.unlink()
                            
                        shutil.copy(ruta_origen_transcripcion, ruta_destino_transcripcion)
                        os.utime(ruta_destino_transcripcion, None)
                        archivos_creados.append(nombre_archivo_transcripcion)
                    else:
                        self.registrar_log(f"  ⚠️ Transcripción no encontrada en: {ruta_origen_transcripcion}")
                else:
                    self.registrar_log(f"  ⚠️ No se encontró audio CALL para {nombre} (DNI: {self.limpiar_id(dni)}, TEL: {self.limpiar_id(telefono)}) - Excel creado")
            else:
                self.registrar_log(f"  ℹ️ consolidados.xlsx no proporcionado - Solo Excel CALL creado para {nombre}")
            
            return True, archivos_creados
            
        except Exception as e:
            self.registrar_log(f"  ❌ Error creando evidencia CALL: {str(e)}")
            return False, archivos_creados
    
    def procesar_cliente(self, fila_cliente: pd.Series, df_nuevos_datos: pd.DataFrame,
                        df_sms: Optional[pd.DataFrame], df_consolidados: Optional[pd.DataFrame],
                        ruta_audio_ivr: str, carpeta_salida_base: Path) -> bool:
        """
        Procesa un cliente individual y crea sus archivos de evidencia
        
        Returns:
            True si se procesó exitosamente
        """
        try:
            # Extraer datos del cliente (limpiando IDs inmediatamente)
            cuenta = self.limpiar_id(fila_cliente['cuenta'])
            nombre = str(fila_cliente['nombre'])
            dni = self.limpiar_id(fila_cliente.get('dni', ''))
            telefono = self.limpiar_id(fila_cliente.get('telefono', ''))
            gestion_efectiva_str = str(fila_cliente['gestion_efectiva'])
            
            # Parsear gestiones efectivas
            gestiones = self.analizar_gestion_efectiva(gestion_efectiva_str)
            
            if not gestiones:
                self.registrar_log(f"⚠️ Cliente {nombre} no tiene gestiones efectivas")
                return False
            
            # Crear carpeta del cliente
            nombre_carpeta = f"{nombre}_{cuenta}"
            carpeta_cliente = carpeta_salida_base / nombre_carpeta
            carpeta_cliente.mkdir(parents=True, exist_ok=True)
            os.utime(carpeta_cliente, None) # Asegurar fecha de modificación actual
            
            self.registrar_log(f"\n📁 Procesando: {nombre_carpeta}")
            self.registrar_log(f"  Gestiones: {', '.join(gestiones)}")
            
            datos_cliente = {
                'cuenta': cuenta,
                'nombre': nombre,
                'dni': dni,
                'telefono': telefono
            }
            
            total_archivos_creados = []
            
            # Procesar IVR
            if 'IVR' in gestiones:
                exito, archivos = self.crear_evidencia_ivr(
                    datos_cliente, df_nuevos_datos, carpeta_cliente, ruta_audio_ivr
                )
                if exito:
                    total_archivos_creados.extend(archivos)
                    self.registrar_log(f"  ✅ IVR: {', '.join(archivos)}")
            
            # Procesar SMS
            if 'SMS' in gestiones and df_sms is not None:
                exito, archivos = self.crear_evidencia_sms(
                    datos_cliente, df_sms, carpeta_cliente
                )
                if exito:
                    total_archivos_creados.extend(archivos)
                    self.registrar_log(f"  ✅ SMS: {', '.join(archivos)}")
            
            # Procesar CALL (df_consolidados es opcional)
            if 'CALL' in gestiones:
                exito, archivos = self.crear_evidencia_call(
                    datos_cliente, df_nuevos_datos, df_consolidados, carpeta_cliente
                )
                if exito:
                    total_archivos_creados.extend(archivos)
                    self.registrar_log(f"  ✅ CALL: {', '.join(archivos)}")
            
            self.registrar_log(f"  📊 Total archivos creados: {len(total_archivos_creados)}")
            
            return True
            
        except Exception as e:
            self.registrar_log(f"❌ Error procesando cliente {nombre}: {str(e)}")
            return False
    
    def validar_campos_dataframe(self, df: pd.DataFrame, campos_requeridos: List[str], 
                                  nombre_archivo: str) -> Tuple[bool, str]:
        """
        Valida que un DataFrame contenga los campos requeridos
        
        Returns:
            Tuple (valido, mensaje_error)
        """
        campos_faltantes = []
        for campo in campos_requeridos:
            if campo not in df.columns:
                campos_faltantes.append(campo)
        
        if campos_faltantes:
            return False, f"{nombre_archivo}: Faltan campos {', '.join(campos_faltantes)}"
        
        return True, ""
